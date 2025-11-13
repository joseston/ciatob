#!/usr/bin/env python3
"""
Script para eliminar hojas de Excel que contengan "REPORTE" o "CONSULTAS"
y guardar el resultado como FARMACIA_LAB.xlsx
"""

from openpyxl import load_workbook
import os
import re

# Ruta del archivo de entrada
archivo_path = r"C:\Users\JOSE\Escritorio\MMC\CIATOB\postgre\migracion\extraccion\CIATOB AGENDA DIARIA 2025.xlsx"

# Nombre del archivo de salida
archivo_salida = r"C:\Users\JOSE\Escritorio\MMC\CIATOB\postgre\migracion\extraccion\laboratorio\FARMACIA_LAB.xlsx"

# Palabras clave para identificar hojas a eliminar
# Se eliminarán las hojas que contengan CUALQUIERA de estas palabras
palabras_clave = ["REPORTE", "CONSULTAS", "CONSULTA"]

print(f"Cargando archivo: {archivo_path}")

# Cargar el archivo Excel
wb = load_workbook(archivo_path)

print(f"\nHojas encontradas en el archivo ({len(wb.sheetnames)} total):")
for idx, sheet_name in enumerate(wb.sheetnames, 1):
    print(f"  {idx}. {sheet_name}")

# Lista para almacenar las hojas que se eliminarán
hojas_a_eliminar = []

# Identificar hojas que contienen las palabras clave
# Se eliminan las hojas que contengan CUALQUIERA de las palabras de la lista
for sheet_name in wb.sheetnames:
    sheet_upper = sheet_name.upper()
    # Verificar si CUALQUIERA de las palabras clave está presente en el nombre de la hoja
    for palabra in palabras_clave:
        if re.search(palabra, sheet_upper):
            hojas_a_eliminar.append(sheet_name)
            break  # Ya encontramos una coincidencia, no necesitamos seguir buscando

print(f"\nHojas que se eliminarán (contienen: {', '.join(palabras_clave)}):")
if hojas_a_eliminar:
    for idx, hoja in enumerate(hojas_a_eliminar, 1):
        print(f"  {idx}. {hoja}")
else:
    print("  Ninguna")

# Eliminar las hojas identificadas
for hoja in hojas_a_eliminar:
    del wb[hoja]
    print(f"✓ Hoja eliminada: {hoja}")

print(f"\nHojas restantes en el archivo:")
for sheet_name in wb.sheetnames:
    print(f"  - {sheet_name}")

# Guardar el archivo modificado
wb.save(archivo_salida)
print(f"\n✓ Archivo guardado exitosamente: {archivo_salida}")
print(f"Total de hojas en el archivo final: {len(wb.sheetnames)}")
